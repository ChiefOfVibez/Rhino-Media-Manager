"""
Populate Excel workbook from products.json
Creates a user-friendly view with previews and hyperlinks
"""
import json
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from pathlib import Path

INPUT_JSON = "products.json"
OUTPUT_EXCEL = r"M:\Proiectare\__SCAN 3D Produse\__BOSCH\__NEW DB__\Bosch_Product_Database.xlsm"

# Column definitions (matching your original layout)
COL_PRODUCT_NAME = 1
COL_DESCRIPTION = 2
COL_PRODUCT_SKU = 3
COL_COD_ARTICOL = 4
COL_RANGE = 5
COL_CATEGORY = 6
COL_SUBCATEGORY = 7
COL_PRODUCT_PREVIEW = 8
COL_OPEN_FOLDER = 9
COL_HOLDER_VARIANT = 10
COL_COLOR_VARIANT = 11
COL_HOLDER_PREVIEW = 12
COL_OPEN_HOLDERS = 13
COL_GH_VARIANT = 14
COL_GH_WIDTH = 15
COL_GH_COLOR = 16
COL_GH_PREVIEW = 17
COL_OPEN_GH = 18
COL_LAST_UPDATED = 19
COL_HAS_PACKAGING = 20
COL_TAGS = 21
COL_NOTES = 22
COL_FOLDER_PATH = 23

COLUMN_NAMES = {
    COL_PRODUCT_NAME: "Product Name",
    COL_DESCRIPTION: "Description",
    COL_PRODUCT_SKU: "SKU",
    COL_COD_ARTICOL: "Cod Articol",
    COL_RANGE: "Range",
    COL_CATEGORY: "Category",
    COL_SUBCATEGORY: "Subcategory",
    COL_PRODUCT_PREVIEW: "Product Preview",
    COL_OPEN_FOLDER: "Open Folder",
    COL_HOLDER_VARIANT: "Holder Variant",
    COL_COLOR_VARIANT: "Color",
    COL_HOLDER_PREVIEW: "Holder Preview",
    COL_OPEN_HOLDERS: "Open Holder",
    COL_GH_VARIANT: "Graphic Holder",
    COL_GH_WIDTH: "Width",
    COL_GH_COLOR: "GH Color",
    COL_GH_PREVIEW: "GH Preview",
    COL_OPEN_GH: "Open Graphic",
    COL_LAST_UPDATED: "Last Updated",
    COL_HAS_PACKAGING: "Packaging",
    COL_TAGS: "Tags",
    COL_NOTES: "Notes",
    COL_FOLDER_PATH: "Folder Path"
}

COLUMN_WIDTHS = {
    COL_PRODUCT_NAME: 25,
    COL_DESCRIPTION: 35,
    COL_PRODUCT_SKU: 20,
    COL_COD_ARTICOL: 20,
    COL_RANGE: 8,
    COL_CATEGORY: 15,
    COL_SUBCATEGORY: 15,
    COL_PRODUCT_PREVIEW: 12,
    COL_OPEN_FOLDER: 12,
    COL_HOLDER_VARIANT: 15,
    COL_COLOR_VARIANT: 12,
    COL_HOLDER_PREVIEW: 12,
    COL_OPEN_HOLDERS: 12,
    COL_GH_VARIANT: 15,
    COL_GH_WIDTH: 8,
    COL_GH_COLOR: 12,
    COL_GH_PREVIEW: 12,
    COL_OPEN_GH: 12,
    COL_LAST_UPDATED: 18,
    COL_HAS_PACKAGING: 10,
    COL_TAGS: 30,
    COL_NOTES: 40,
    COL_FOLDER_PATH: 60
}

def create_or_update_workbook(products):
    """Create or update Excel workbook with products data"""
    output_path = Path(OUTPUT_EXCEL)
    
    # Check if file exists and is locked
    if output_path.exists():
        if is_file_locked(output_path):
            print("\n" + "="*60)
            print("⚠️  WARNING: Excel file is currently OPEN!")
            print("="*60)
            print("Please CLOSE the Excel file and try again.")
            print(f"File: {output_path}")
            print("="*60 + "\n")
            input("Press Enter after closing the file...")
            
            # Check again
            if is_file_locked(output_path):
                raise Exception("File is still open. Cannot update.")
        
        # Create backup
        create_backup(output_path)
        
        print(f"Updating existing file: {output_path}")
        try:
            wb = openpyxl.load_workbook(output_path, keep_vba=True)
            ws = wb.active
            # Clear existing data (keep headers and buttons)
            clear_data_rows(ws)
        except Exception as e:
            print(f"Error loading existing file: {e}")
            print("Creating fresh file instead...")
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Product Database"
            create_buttons(ws)
            create_headers(ws)
    else:
        print(f"Creating new file: {output_path}")
        # Ensure directory exists
        output_path.parent.mkdir(parents=True, exist_ok=True)
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Product Database"
        # Create buttons and headers for new file
        create_buttons(ws)
        create_headers(ws)
    
    # Populate data (starts at row 3, after buttons and headers)
    row = 3
    for product in products:
        populate_row(ws, row, product)
        row += 1
    
    # Set row heights
    ws.row_dimensions[1].height = 30  # Button row
    ws.row_dimensions[2].height = 20  # Header row
    for r in range(3, row):
        ws.row_dimensions[r].height = 60  # Data rows (for preview images)
    
    # Freeze panes (first 2 rows)
    ws.freeze_panes = "A3"
    
    return wb

def is_file_locked(filepath):
    """Check if file is locked (open in another program)"""
    try:
        # Try to open with write access
        with open(filepath, 'a'):
            pass
        return False
    except IOError:
        return True

def create_backup(filepath):
    """Create backup of existing file"""
    import shutil
    from datetime import datetime
    
    backup_name = filepath.stem + "_backup_" + datetime.now().strftime("%Y%m%d_%H%M%S") + filepath.suffix
    backup_path = filepath.parent / backup_name
    
    try:
        shutil.copy2(filepath, backup_path)
        print(f"✓ Backup created: {backup_path.name}")
    except Exception as e:
        print(f"Warning: Could not create backup: {e}")

def create_buttons(ws):
    """Create action buttons in row 1"""
    # Button row styling
    button_fill = PatternFill(start_color="4CAF50", end_color="4CAF50", fill_type="solid")
    button_font = Font(bold=True, color="FFFFFF", size=11)
    button_alignment = Alignment(horizontal="center", vertical="center")
    
    # Scan Database button
    ws.merge_cells('A1:C1')
    btn1 = ws['A1']
    btn1.value = "🔄 SCAN DATABASE"
    btn1.font = button_font
    btn1.fill = button_fill
    btn1.alignment = button_alignment
    
    # Refresh Excel button
    ws.merge_cells('D1:F1')
    btn2 = ws['D1']
    btn2.value = "📊 REFRESH EXCEL"
    btn2.font = button_font
    btn2.fill = button_fill
    btn2.alignment = button_alignment
    
    # Insert Previews button
    ws.merge_cells('G1:I1')
    btn3 = ws['G1']
    btn3.value = "🖼️ INSERT PREVIEWS"
    btn3.font = button_font
    btn3.fill = button_fill
    btn3.alignment = button_alignment
    
    # Export Notes button
    ws.merge_cells('J1:L1')
    btn4 = ws['J1']
    btn4.value = "💾 EXPORT NOTES"
    btn4.font = button_font
    btn4.fill = button_fill
    btn4.alignment = button_alignment

def clear_data_rows(ws):
    """Clear data rows but keep headers and buttons"""
    # Delete all rows from 3 onwards
    if ws.max_row > 2:
        ws.delete_rows(3, ws.max_row - 2)

def create_headers(ws):
    """Create header row with styling (row 2)"""
    # Header row (row 2)
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    
    for col_num, col_name in COLUMN_NAMES.items():
        cell = ws.cell(row=2, column=col_num)
        cell.value = col_name
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        
        # Set column width
        col_letter = get_column_letter(col_num)
        ws.column_dimensions[col_letter].width = COLUMN_WIDTHS.get(col_num, 15)
    
    # Hide folder path column (but keep the data)
    ws.column_dimensions[get_column_letter(COL_FOLDER_PATH)].hidden = True

def populate_row(ws, row, product):
    """Populate a single row with product data"""
    # Basic info
    ws.cell(row, COL_PRODUCT_NAME).value = product.get('productName', '')
    ws.cell(row, COL_DESCRIPTION).value = product.get('description', '')
    ws.cell(row, COL_PRODUCT_SKU).value = product.get('sku', '')
    ws.cell(row, COL_COD_ARTICOL).value = product.get('codArticol', '')
    ws.cell(row, COL_RANGE).value = product.get('range', '')
    ws.cell(row, COL_CATEGORY).value = product.get('category', '')
    ws.cell(row, COL_SUBCATEGORY).value = product.get('subcategory', '')
    
    # Preview columns (display text for now, VBA will handle images)
    preview_path = product.get('productPreview', '')
    if preview_path:
        cell = ws.cell(row, COL_PRODUCT_PREVIEW)
        cell.value = "Preview"
        cell.hyperlink = f"file:///{preview_path}"
        cell.font = Font(color="0563C1", underline="single")
    
    # Open folder link
    folder_path = product.get('folderPath', '')
    if folder_path:
        cell = ws.cell(row, COL_OPEN_FOLDER)
        cell.value = "Open Folder"
        cell.hyperlink = f"file:///{folder_path}"
        cell.font = Font(color="0563C1", underline="single")
    
    # Holder info
    ws.cell(row, COL_HOLDER_VARIANT).value = product.get('holderVariant', '')
    ws.cell(row, COL_COLOR_VARIANT).value = product.get('colorVariant', '')
    
    # Holder preview
    holder_preview = product.get('holderPreview', '')
    if holder_preview:
        cell = ws.cell(row, COL_HOLDER_PREVIEW)
        cell.value = "Preview"
        cell.hyperlink = f"file:///{holder_preview}"
        cell.font = Font(color="0563C1", underline="single")
    
    # Open holder .3dm link
    if product.get('holders'):
        holder_variant = product.get('holderVariant', '')
        color_variant = product.get('colorVariant', '')
        if holder_variant and color_variant:
            # Find the .3dm file path
            for holder in product['holders']:
                if holder['variant'] == holder_variant:
                    for file_path in holder.get('files', []):
                        if color_variant in file_path:
                            cell = ws.cell(row, COL_OPEN_HOLDERS)
                            cell.value = "Open .3dm"
                            cell.hyperlink = f"file:///{file_path}"
                            cell.font = Font(color="0563C1", underline="single")
                            break
    
    # Graphic holder info
    ws.cell(row, COL_GH_VARIANT).value = product.get('ghVariant', '')
    ws.cell(row, COL_GH_WIDTH).value = product.get('ghWidth', '')
    ws.cell(row, COL_GH_COLOR).value = product.get('ghColor', '')
    
    # Graphic preview (placeholder for now)
    cell = ws.cell(row, COL_GH_PREVIEW)
    cell.value = ""  # Can be enhanced
    
    # Open graphic .3dm link
    if product.get('graphics'):
        gh_variant = product.get('ghVariant', '')
        gh_width = product.get('ghWidth', '')
        gh_color = product.get('ghColor', '')
        if gh_variant and gh_width and gh_color:
            for graphic in product['graphics']:
                if graphic['prefix'] == gh_variant and str(graphic['width']) == str(gh_width):
                    for file_path in graphic.get('files', []):
                        if gh_color in file_path:
                            cell = ws.cell(row, COL_OPEN_GH)
                            cell.value = "Open .3dm"
                            cell.hyperlink = f"file:///{file_path}"
                            cell.font = Font(color="0563C1", underline="single")
                            break
    
    # Metadata
    ws.cell(row, COL_LAST_UPDATED).value = product.get('lastUpdated', '')
    ws.cell(row, COL_HAS_PACKAGING).value = "Yes" if product.get('hasPackaging') else "No"
    ws.cell(row, COL_TAGS).value = ", ".join(product.get('tags', []))
    ws.cell(row, COL_NOTES).value = product.get('notes', '')
    ws.cell(row, COL_FOLDER_PATH).value = folder_path
    
    # Center align some columns
    for col in [COL_RANGE, COL_PRODUCT_PREVIEW, COL_OPEN_FOLDER, 
                COL_HOLDER_PREVIEW, COL_OPEN_HOLDERS, COL_GH_WIDTH, 
                COL_GH_PREVIEW, COL_OPEN_GH, COL_HAS_PACKAGING]:
        ws.cell(row, col).alignment = Alignment(horizontal="center", vertical="center")
    
    # Wrap text for description and notes
    for col in [COL_DESCRIPTION, COL_NOTES]:
        ws.cell(row, col).alignment = Alignment(wrap_text=True, vertical="top")

def main():
    """Main entry point"""
    print("=" * 60)
    print("Excel Workbook Generator")
    print("=" * 60)
    print()
    
    # Load JSON
    json_path = Path(INPUT_JSON)
    if not json_path.exists():
        print(f"ERROR: {INPUT_JSON} not found!")
        print("Please run bosch_scanner.py first to generate the JSON file.")
        return
    
    print(f"Loading: {json_path}")
    with open(json_path, 'r', encoding='utf-8') as f:
        products = json.load(f)
    
    print(f"Found {len(products)} products")
    
    # Create or update workbook
    print("\nProcessing products...")
    wb = create_or_update_workbook(products)
    
    # Save
    output_path = Path(OUTPUT_EXCEL)
    print(f"\nSaving to: {output_path}")
    
    try:
        wb.save(output_path)
        print("\n" + "="*60)
        print("✅ SUCCESS!")
        print("="*60)
        print(f"Excel file updated: {output_path.name}")
        print(f"Location: {output_path.parent}")
        print(f"Total products: {len(products)}")
        print("\nNext steps:")
        print("  1. Open the Excel file")
        print("  2. Click 'INSERT PREVIEWS' button to show images")
        print("  3. Add notes if needed")
        print("="*60)
    except PermissionError:
        print("\n" + "="*60)
        print("❌ ERROR: Permission denied!")
        print("="*60)
        print("The file might be:")
        print("  - Open in Excel (close it first)")
        print("  - Read-only (check file permissions)")
        print("  - On a network drive with no write access")
        print(f"\nFile: {output_path}")
        print("="*60)
        raise
    except Exception as e:
        print(f"\n❌ ERROR saving file: {e}")
        raise

if __name__ == "__main__":
    try:
        main()
    except Exception as e:
        print(f"ERROR: {e}")
        import traceback
        traceback.print_exc()
    
    input("Press Enter to exit...")
