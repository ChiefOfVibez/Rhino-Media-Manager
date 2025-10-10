"""
Populate Excel workbook from JSON files directly
Reads product JSONs and creates Excel with dropdowns and links
"""
import json
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from pathlib import Path

# Paths
BASE_PATH = r"\\MattHQ-SVDC01\Share\Proiectare\__SCAN 3D Produse\__BOSCH\__NEW DB__"
OUTPUT_EXCEL = r"M:\Proiectare\__SCAN 3D Produse\__BOSCH\__NEW DB__\Bosch_Product_Database.xlsm"
OUTPUT_EXCEL_XLSX = r"M:\Proiectare\__SCAN 3D Produse\__BOSCH\__NEW DB__\Bosch_Product_Database.xlsx"

# Column definitions (simplified - removed graphic holder columns)
COL_PRODUCT_NAME = 1    # A
COL_DESCRIPTION = 2     # B
COL_PRODUCT_SKU = 3     # C
COL_COD_ARTICOL = 4     # D - Formula based on J+K
COL_RANGE = 5           # E
COL_CATEGORY = 6        # F
COL_SUBCATEGORY = 7     # G
COL_PRODUCT_PREVIEW = 8 # H - Clickable image link
COL_OPEN_FOLDER = 9     # I
COL_HOLDER_VARIANT = 10 # J - Dropdown
COL_COLOR_VARIANT = 11  # K - Dropdown
COL_HOLDER_PREVIEW = 12 # L - Live updated
COL_OPEN_HOLDER = 13    # M - Reveal in Explorer
COL_LAST_UPDATED = 14   # N
COL_HAS_PACKAGING = 15  # O
COL_TAGS = 16           # P
COL_NOTES = 17          # Q
COL_FOLDER_PATH = 18    # R - Hidden

COLUMN_NAMES = {
    COL_PRODUCT_NAME: "Product Name",
    COL_DESCRIPTION: "Description",
    COL_PRODUCT_SKU: "SKU",
    COL_COD_ARTICOL: "Cod Articol",
    COL_RANGE: "Range",
    COL_CATEGORY: "Category",
    COL_SUBCATEGORY: "Subcategory",
    COL_PRODUCT_PREVIEW: "Product Preview",
    COL_OPEN_FOLDER: "Open Folder",
    COL_HOLDER_VARIANT: "Holder Variant",
    COL_COLOR_VARIANT: "Color",
    COL_HOLDER_PREVIEW: "Holder Preview",
    COL_OPEN_HOLDER: "Reveal in Explorer",
    COL_LAST_UPDATED: "Last Updated",
    COL_HAS_PACKAGING: "Packaging",
    COL_TAGS: "Tags",
    COL_NOTES: "Notes",
    COL_FOLDER_PATH: "Folder Path"
}

COLUMN_WIDTHS = {
    COL_PRODUCT_NAME: 25,
    COL_DESCRIPTION: 40,
    COL_PRODUCT_SKU: 20,
    COL_COD_ARTICOL: 20,
    COL_RANGE: 8,
    COL_CATEGORY: 15,
    COL_SUBCATEGORY: 15,
    COL_PRODUCT_PREVIEW: 15,
    COL_OPEN_FOLDER: 12,
    COL_HOLDER_VARIANT: 15,
    COL_COLOR_VARIANT: 12,
    COL_HOLDER_PREVIEW: 15,
    COL_OPEN_HOLDER: 12,
    COL_LAST_UPDATED: 18,
    COL_HAS_PACKAGING: 10,
    COL_TAGS: 35,
    COL_NOTES: 40,
    COL_FOLDER_PATH: 60
}

def scan_json_files():
    """Scan all JSON files in product folders"""
    products = []
    tools_path = Path(BASE_PATH) / "Tools and Holders"
    
    print(f"Scanning: {tools_path}")
    
    # Walk through ranges (PRO/DIY)
    for range_folder in sorted(tools_path.iterdir()):
        if not range_folder.is_dir() or range_folder.name.startswith('_') or range_folder.name == 'Holders':
            continue
        
        range_name = range_folder.name
        print(f"  Range: {range_name}")
        
        # Walk through categories
        for category_folder in sorted(range_folder.iterdir()):
            if not category_folder.is_dir() or category_folder.name.startswith('_'):
                continue
            
            category_name = category_folder.name
            print(f"    Category: {category_name}")
            
            # Walk through products
            for product_folder in sorted(category_folder.iterdir()):
                if not product_folder.is_dir() or product_folder.name.startswith('_'):
                    continue
                
                # Find JSON file
                json_files = list(product_folder.glob('*.json'))
                if not json_files:
                    continue
                
                json_file = json_files[0]
                try:
                    with open(json_file, 'r', encoding='utf-8') as f:
                        product_data = json.load(f)
                        product_data['_folderPath'] = str(product_folder)
                        product_data['_jsonPath'] = str(json_file)
                        products.append(product_data)
                        print(f"      ✓ {product_data.get('productName', 'Unknown')}")
                except Exception as e:
                    print(f"      ✗ Error reading {json_file.name}: {e}")
    
    return products

def create_or_update_workbook(products):
    """Create or update Excel workbook"""
    # Save as .xlsx for maximum compatibility, then user can Save As .xlsm from Excel
    output_path = Path(OUTPUT_EXCEL_XLSX)
    
    # Check if file is locked
    if output_path.exists() and is_file_locked(output_path):
        print("\n" + "="*60)
        print("⚠️  Excel file is OPEN! Please close it first.")
        print("="*60)
        input("Press Enter after closing...")
        # Check again and abort if still open
        if is_file_locked(output_path):
            raise Exception("File is still open!")
    
    # Create or load workbook
    if output_path.exists():
        print(f"Updating: {output_path}")
        wb = openpyxl.load_workbook(output_path, keep_vba=False)
        ws = wb.active
        clear_data_rows(ws)
    else:
        print(f"Creating: {output_path}")
        output_path.parent.mkdir(parents=True, exist_ok=True)
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Product Database"
        create_buttons(ws)
        create_headers(ws)
    
    # Populate data
    row = 3
    for product in products:
        populate_row(ws, row, product)
        row += 1
    
    # Set row heights
    ws.row_dimensions[1].height = 30
    ws.row_dimensions[2].height = 20
    for r in range(3, row):
        ws.row_dimensions[r].height = 60
    
    # Freeze panes
    ws.freeze_panes = "A3"
    
    return wb

def is_file_locked(filepath):
    """Check if file is locked"""
    try:
        with open(filepath, 'a'):
            pass
        return False
    except IOError:
        return True

def create_buttons(ws):
    """Create action buttons in row 1"""
    button_fill = PatternFill(start_color="4CAF50", end_color="4CAF50", fill_type="solid")
    button_font = Font(bold=True, color="FFFFFF", size=11)
    button_alignment = Alignment(horizontal="center", vertical="center")
    
    ws.merge_cells('A1:C1')
    ws['A1'].value = "🔄 SCAN DATABASE"
    ws['A1'].font = button_font
    ws['A1'].fill = button_fill
    ws['A1'].alignment = button_alignment
    
    ws.merge_cells('D1:F1')
    ws['D1'].value = "📊 REFRESH EXCEL"
    ws['D1'].font = button_font
    ws['D1'].fill = button_fill
    ws['D1'].alignment = button_alignment
    
    ws.merge_cells('G1:I1')
    ws['G1'].value = "🖼️ INSERT PREVIEWS"
    ws['G1'].font = button_font
    ws['G1'].fill = button_fill
    ws['G1'].alignment = button_alignment
    
    ws.merge_cells('J1:L1')
    ws['J1'].value = "💾 EXPORT NOTES"
    ws['J1'].font = button_font
    ws['J1'].fill = button_fill
    ws['J1'].alignment = button_alignment

def clear_data_rows(ws):
    """Clear data rows but keep headers"""
    if ws.max_row > 2:
        ws.delete_rows(3, ws.max_row - 2)

def create_headers(ws):
    """Create header row"""
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    
    for col_num, col_name in COLUMN_NAMES.items():
        cell = ws.cell(row=2, column=col_num)
        cell.value = col_name
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        
        col_letter = get_column_letter(col_num)
        ws.column_dimensions[col_letter].width = COLUMN_WIDTHS.get(col_num, 15)
    
    # Hide folder path column
    ws.column_dimensions[get_column_letter(COL_FOLDER_PATH)].hidden = True

def populate_row(ws, row, product):
    """Populate a single row with product data"""
    folder_path = product.get('_folderPath', '')
    
    # Basic info
    ws.cell(row, COL_PRODUCT_NAME).value = product.get('productName', '')
    ws.cell(row, COL_DESCRIPTION).value = product.get('description', '')
    ws.cell(row, COL_PRODUCT_SKU).value = product.get('sku', '')
    ws.cell(row, COL_RANGE).value = product.get('range', '')
    ws.cell(row, COL_CATEGORY).value = product.get('category', '')
    ws.cell(row, COL_SUBCATEGORY).value = product.get('subcategory', '')
    
    # Formula for Cod Articol (lookup based on variant+color from holders array)
    # For now, leave empty - will be filled by VBA when dropdown changes
    ws.cell(row, COL_COD_ARTICOL).value = ""
    
    # Product preview - clickable link
    previews = product.get('previews', {})
    mesh_preview = previews.get('meshPreview', {})
    
    # Handle both dict format (with fullPath) and string format
    if isinstance(mesh_preview, dict):
        preview_file = mesh_preview.get('fullPath', '')
    elif isinstance(mesh_preview, str):
        preview_file = folder_path + "\\" + mesh_preview
    else:
        preview_file = ""
    
    if preview_file:
        cell = ws.cell(row, COL_PRODUCT_PREVIEW)
        cell.value = "🖼️ View"
        # Store path as formula to make it clickable
        cell.value = '=HYPERLINK("' + preview_file + '", "🖼️ View")'
        cell.font = Font(color="0563C1", underline="single")
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # Open folder link
    if folder_path:
        cell = ws.cell(row, COL_OPEN_FOLDER)
        cell.value = '=HYPERLINK("' + folder_path + '", "📁 Open")'
        cell.font = Font(color="0563C1", underline="single")
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # Holder dropdowns
    holders = product.get('holders', [])
    if holders:
        # Get unique variants and colors
        variants = sorted(set(h.get('variant', '') for h in holders if h.get('variant')))
        colors = sorted(set(h.get('color', '') for h in holders if h.get('color')))
        
        if variants and colors:
            # Set default (first holder)
            ws.cell(row, COL_HOLDER_VARIANT).value = holders[0].get('variant', '')
            ws.cell(row, COL_COLOR_VARIANT).value = holders[0].get('color', '')
            
            # Create dropdown for variant
            variant_list = ",".join(variants)
            dv_variant = DataValidation(type="list", formula1=f'"{variant_list}"', allow_blank=False)
            ws.add_data_validation(dv_variant)
            dv_variant.add(ws.cell(row, COL_HOLDER_VARIANT))
            
            # Create dropdown for color
            color_list = ",".join(colors)
            dv_color = DataValidation(type="list", formula1=f'"{color_list}"', allow_blank=False)
            ws.add_data_validation(dv_color)
            dv_color.add(ws.cell(row, COL_COLOR_VARIANT))
            
            # Set initial holder preview and link
            update_holder_info(ws, row, holders, holders[0].get('variant', ''), holders[0].get('color', ''))
    
    # Tags
    tags = product.get('tags', [])
    ws.cell(row, COL_TAGS).value = ", ".join(tags) if tags else ""
    
    # Metadata
    metadata = product.get('metadata', {})
    ws.cell(row, COL_LAST_UPDATED).value = metadata.get('lastModified', '')
    
    # Packaging
    packaging = product.get('packaging', {})
    ws.cell(row, COL_HAS_PACKAGING).value = "Yes" if packaging.get('fileName') else "No"
    ws.cell(row, COL_HAS_PACKAGING).alignment = Alignment(horizontal="center", vertical="center")
    
    # Notes
    ws.cell(row, COL_NOTES).value = product.get('notes', '')
    
    # Hidden folder path
    ws.cell(row, COL_FOLDER_PATH).value = folder_path
    
    # Wrap text for description and notes
    ws.cell(row, COL_DESCRIPTION).alignment = Alignment(wrap_text=True, vertical="top")
    ws.cell(row, COL_NOTES).alignment = Alignment(wrap_text=True, vertical="top")
    ws.cell(row, COL_TAGS).alignment = Alignment(wrap_text=True, vertical="top")

def update_holder_info(ws, row, holders, variant, color):
    """Update holder preview and link based on selection"""
    # Find matching holder
    for holder in holders:
        if holder.get('variant') == variant and holder.get('color') == color:
            # Set Cod Articol
            ws.cell(row, COL_COD_ARTICOL).value = holder.get('codArticol', '')
            
            # Holder preview link - only create if full path exists
            preview = holder.get('preview', '')
            if preview:
                # If preview is a dict with fullPath, use that
                if isinstance(preview, dict):
                    preview_path = preview.get('fullPath', '')
                else:
                    preview_path = preview
                
                # Only create link if it's a full path (has \\ or starts with drive letter)
                if preview_path and ('\\\\' in preview_path or ':' in preview_path):
                    cell = ws.cell(row, COL_HOLDER_PREVIEW)
                    cell.value = f'=HYPERLINK("{preview_path}", " Reveal in Explorer")'
                    cell.font = Font(color="0563C1", underline="single")
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                else:
                    # Preview file not found - leave empty
                    ws.cell(row, COL_HOLDER_PREVIEW).value = ""
            
            # Holder reveal in explorer - open containing folder
            full_path = holder.get('fullPath', '')
            if full_path:
                # Extract folder path from full file path
                import os
                folder = os.path.dirname(full_path)
                cell = ws.cell(row, COL_OPEN_HOLDER)
                cell.value = '=HYPERLINK("' + folder + '", "📁 Reveal")'
                cell.font = Font(color="0563C1", underline="single")
                cell.alignment = Alignment(horizontal="center", vertical="center")
            break

def main():
    """Main entry point"""
    print("="*70)
    print("Excel Database Generator - Direct JSON Reader")
    print("="*70)
    print()
    
    # Scan JSON files
    products = scan_json_files()
    
    if not products:
        print("\n⚠️  No products found!")
        print("Check that JSON files exist in product folders.")
        return
    
    print(f"\nFound {len(products)} products")
    print()
    
    # Create/update Excel
    wb = create_or_update_workbook(products)
    
    # Save
    print(f"\nSaving to: {OUTPUT_EXCEL_XLSX}")
    try:
        # Ensure directory exists, then save .xlsx
        Path(OUTPUT_EXCEL_XLSX).parent.mkdir(parents=True, exist_ok=True)
        wb.save(OUTPUT_EXCEL_XLSX)
        print("✓ Excel file updated successfully!")
    except Exception as e:
        print(f"✗ Error saving Excel: {e}")
        return
    
    print()
    print("="*70)
    print("✅ Complete!")
    print("="*70)
    print()
    print(f"Open Excel: {OUTPUT_EXCEL_XLSX}")
    print()

if __name__ == "__main__":
    try:
        main()
    except Exception as e:
        print(f"\n❌ ERROR: {e}")
        import traceback
        traceback.print_exc()
    
    input("\nPress Enter to exit...")
